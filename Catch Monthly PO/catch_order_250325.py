import cx_Oracle
import pandas as pd
import re
import os
import argparse
from pypdf import PdfWriter
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime


parser = argparse.ArgumentParser(description="Generate Monthly Order List from ERP")
args = parser.parse_args()


def ERP_Connection(SQL):
    # Define connection details
    dsn = cx_Oracle.makedsn(
        host="192.168.1.242",      # Replace with your host IP or domain
        port=1526,                 # Replace with your port
        service_name="sperpdb"      # Replace with your service name
    )

    # Establish the connection
    connection = cx_Oracle.connect(
        user="spselect",          # Replace with your username
        password="select",        # Replace with your password
        dsn=dsn
    )

    # Execute SQL query and return dataframe
    df = pd.read_sql_query(SQL, connection)

    # Close connection
    connection.close()
    
    return df


def Create_Monthly_Order_List(SC_START, SQL):

    Monthly_Order = ERP_Connection(SQL).copy()

    # Remove duplicate rows according to SC_NO
    if "SC_NO" in Monthly_Order.columns:
        Monthly_Order = Monthly_Order.drop_duplicates(subset=["SC_NO"]).sort_values(by="SC_NO")

    Monthly_Order["SC_NO"] = pd.to_numeric(Monthly_Order["SC_NO"], errors="coerce")

    new_rows = []
    
    # Iterate over SC_NO to find missing numbers
    for i in range(len(Monthly_Order)-1): #last SC_NO doesn't have to be in the for loop cuz there's no next_sc
        current_sc = Monthly_Order.iloc[i]["SC_NO"]
        next_sc = Monthly_Order.iloc[i + 1]["SC_NO"]
        
        # If there is a gap in the sequence, add a missing row
        if next_sc != current_sc + 1:
            missing_sc = current_sc + 1  # The missing SC_NO
            new_rows.append({"SC_NO": missing_sc, "CST_REFE_NO": "訂單刪除", "ORD_CST_NO": ""})

    if new_rows:
        Monthly_Order = pd.concat([Monthly_Order, pd.DataFrame(new_rows)])

    Monthly_Order = Monthly_Order.sort_values(by="SC_NO").reset_index(drop=True)


    date_obj = datetime.strptime(SC_START, "%y%m")
    MONTH = date_obj.strftime("%Y.%m月")

    Arrange_File = fr"Z:\業務部\業務一課\H-訂單\5. 每月訂單\2025\{MONTH}份訂單-Caysie.xlsx"
    wb = Workbook()
    ws = wb.active

    #Fill informations
    ws["A2"] = "Dear 總經理,"

    ws["A4"] = MONTH

    ws["A5"] = \
    f"訂單號碼 {Monthly_Order.loc[0, 'SC_NO']} ~ {Monthly_Order.loc[Monthly_Order.shape[0]-1, 'SC_NO']}"

    # Function to check if a string contains Chinese characters
    def contains_chinese(text):
        return bool(re.search(r'[\u4e00-\u9fff]', str(text)))

    Special_Orders = Monthly_Order[Monthly_Order["CST_REFE_NO"].apply(contains_chinese)]

    #for rows then for the column in the roll to get the cell
    for r_idx, row in enumerate(Special_Orders.itertuples(index=False), start=11): #start row 11
        for c_idx, value in enumerate(row, start=2): # start column 2
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Order amount = all orders-special orders(庫存單、樣品單、已刪除訂單)
    ws["A6"] = f"共 {Monthly_Order.shape[0]-Special_Orders.shape[0]} 張訂單"

    ws["A7"] = "謝謝"

    #column B width for SC numbers, column C width for special order names
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30

    # Set font name and size
    font_style = Font(name="Calibri", size=18) 

    # Apply the font to all cells in the sheet
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font_style

    # Create a new sheet for the Monthly Order List
    ws_list = wb.create_sheet(title="Monthly Order List")

    # Write Monthly_Order DataFrame into the new sheet
    for r_idx, row in enumerate(Monthly_Order.itertuples(index=False), start=1):  # Start at row 1
        for c_idx, value in enumerate(row, start=1):  # Start at column 1
            ws_list.cell(row=r_idx, column=c_idx, value=value)


    wb.save(Arrange_File)
    
    return Arrange_File

def Save_Merged_Confirmed_POpdfs(SC_START):
    Order_Directory = r"Z:\業務部\業務一課\H-訂單\1. 外銷"
    
    date_obj = datetime.strptime(SC_START, "%y%m")
    month = date_obj.strftime("%Y%m")

    PO_pdfs = {}
    for root, _, files in os.walk(Order_Directory):
            if month in os.path.basename(root):
                    latest_pdf = None
                    latest_time = 0 
                    for file in files:
                        if file.lower().endswith(".pdf"):  # Filter PDF files
                            file_path = os.path.join(root, file)
                            file_mtime = os.path.getmtime(file_path)  # Get modification time
                        
                            if file_mtime > latest_time : # Update if it's the latest file
                                latest_time = file_mtime
                                latest_pdf = file_path
                        
                            # Print the latest updated PDF file
                            if latest_pdf and "出貨" not in root: 
                                PO_pdfs[root] = (latest_pdf, latest_time)

    PO_pdfs = sorted(PO_pdfs.values(), key=lambda x: x[1], reverse=False)                  

    if PO_pdfs:
        writer = PdfWriter()
        for pdf, _ in PO_pdfs:  # Ignore modification time while appending
            writer.append(pdf)  # Append each latest PDF
        
        output_pdf_path = os.path.join(r"Z:\業務部\業務一課\H-訂單\5. 每月訂單\2025\確認訂單", f"PrintPO_{month}.pdf")
        with open(output_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)
        
        print(f"Merged PDF saved: {output_pdf_path}")
    else:
        print("No PDFs found to merge.")




if __name__ == "__main__":
    # Get user input
    SC_START = input("SC start: ")
    SQL = (f"SELECT SC_NO, CST_REFE_NO, ORD_CST_NO FROM V_SCH0200Q_ORD WHERE SC_NO LIKE '{SC_START}%'")

    # Get data from ERP
    connection = ERP_Connection(SQL)
    order = Create_Monthly_Order_List(SC_START, SQL)  # Pass the correct query
    pdfs = Save_Merged_Confirmed_POpdfs(SC_START)

    # Save confirmation message
    print(f"List Saved Successfully in Z:/業務部/業務一課/H-訂單/5. 每月訂單/2025")


