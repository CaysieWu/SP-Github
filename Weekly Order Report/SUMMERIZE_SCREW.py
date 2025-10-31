import pandas as pd
from sqlalchemy import create_engine
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import sqlite3
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import win32com.client as win32
import time

#以下為外銷接單統計

def GET_DATE():

	today = datetime.now().strftime('%Y/%m/%d')
	year = datetime.now().year
	last_year = datetime.now()-relativedelta(years=1)
	last_year = last_year.year
	next_year = datetime.now()+relativedelta(years=1)
	next_year = next_year.year

	return year, last_year, today, next_year

def ERP_CONNECT(year, last_year):

	oracle_connection_string = (
		"oracle+cx_oracle://spselect:select@192.168.1.242:1526/?service_name=sperpdb"
	)
	engine = create_engine(oracle_connection_string)

	# query1: 找出庫存單的SC以利後續排除接單及出貨計算
	query1 = f"""
		SELECT SC_NO FROM V_SCH0200Q_ORD
		WHERE CST_REFE_NO LIKE '%庫存單%'
	"""
	stock_order = pd.read_sql_query(query1, engine.connect())
	stock_order = "', '".join(stock_order["sc_no"].astype(str))
   
	# query2: 取SC及客戶代號，分為確認及未確認訂單
	query2 = f"""
		SELECT SC_NO, ORD_CST_NO, CONFIRM_DATE FROM ssl_cst_orde_m
		WHERE ORD_DATE >= TO_DATE('{last_year}-12-01', 'YYYY-MM-DD') 
		AND ORD_DATE <= TO_DATE('{year}-12-31', 'YYYY-MM-DD')
		AND SC_NO NOT IN ('{stock_order}')
		AND END_CODE != 'D'
	"""
	sc = pd.read_sql_query(query2, engine.connect())
	sc.columns = sc.columns.str.upper()
	confirm_sc = sc[sc['CONFIRM_DATE'].notna()]
	unconfirm_sc = sc[sc['CONFIRM_DATE'].isna()]
	confirm_scs = "', '".join(confirm_sc["SC_NO"].astype(str))
	unconfirm_scs = "', '".join(unconfirm_sc["SC_NO"].astype(str))

	# query3: 未確認訂單，用於後續統計重量
	query3 = f"""
		SELECT SC_NO, ORDER_WEIG FROM ssl_cst_orde_d
		WHERE SC_NO IN ('{unconfirm_scs}')
		AND SC_NO NOT IN ('{stock_order}')
		AND END_CODE != 'D'
	"""
	unconfirm = pd.read_sql_query(query3, engine.connect())
	unconfirm.columns = unconfirm.columns.str.upper()
	
	#query4: 取接單重量-只跑上周之已確認訂單-且SC不得為庫存單之SC-且不取用取消項次
	query4 = f"""
		SELECT SC_NO, ORDER_WEIG FROM ssl_cst_orde_d
		WHERE SC_NO IN ('{confirm_scs}')
		AND SC_NO NOT IN ('{stock_order}')
		AND END_CODE != 'D'
	"""
	confirm = pd.read_sql_query(query4, engine.connect())
	confirm.columns = confirm.columns.str.upper()

	# query5: 取訂單交期於本年度之訂單重量-且不得為庫存單SC-且不得為取消項次
	query5 = f"""
		SELECT DLV_DATE, ORDER_WEIG,SC_NO FROM ssl_cst_orde_d
		WHERE DLV_DATE >= TO_DATE('{last_year}-12-01', 'YYYY-MM-DD')
		AND  DLV_DATE < TO_DATE('{next_year}-05-31', 'YYYY-MM-DD')
		AND  SC_NO NOT IN ('{stock_order}')
		AND END_CODE != 'D'
	"""
	ship = pd.read_sql_query(query5, engine.connect())
	ship.columns = ship.columns.str.upper()

	#計算已打單未確認之訂單
	unconfirm['ORDER_WEIG_MT'] = unconfirm['ORDER_WEIG'] / 1000
	unconfirm = unconfirm[["SC_NO", "ORDER_WEIG_MT"]].groupby("SC_NO").sum()

	#用SC合併資料，產出每月接單明細，提取每月資料可免去月底校正
	confirm = confirm[["SC_NO", "ORDER_WEIG"]].groupby("SC_NO").sum()
	confirm_sc["ORDER_WEIG"] = confirm_sc["SC_NO"].map(confirm["ORDER_WEIG"])
	confirm_sc['ORDER_WEIG_MT'] = confirm_sc['ORDER_WEIG'] / 1000
	confirm_sc["YEAR_MONTH"] = confirm_sc["CONFIRM_DATE"].dt.strftime('%Y/%m')
	confirm_sc["CONFIRM_DATE"] = confirm_sc["CONFIRM_DATE"].dt.strftime('%Y/%m/%d')

	monthly_dfs = {
	month: df[["SC_NO", "ORD_CST_NO", "CONFIRM_DATE", "ORDER_WEIG"]].copy()
	for month, df in confirm_sc.groupby("YEAR_MONTH")}

	#計算每月接單總重
	month_summerize = confirm_sc[["YEAR_MONTH", "ORDER_WEIG_MT"]].groupby("YEAR_MONTH").sum()

	#用query5取得本年度每月預計出貨量
	ship['MONTH'] = ship['DLV_DATE'].dt.strftime('%Y/%m')
	ship['ORDER_WEIG_MT'] = ship['ORDER_WEIG'] / 1000
	expect_ship = ship[["MONTH", "ORDER_WEIG_MT"]].groupby("MONTH").sum()
	
	return unconfirm, monthly_dfs, month_summerize, expect_ship

def DB_CONNECT(year, last_year):
	db_path = r"Z:\跨部門\共用資料夾\C. 業務部\詢價統計DB\QUOTATION_DATABASE.db"
	conn = sqlite3.connect(db_path)
	
	#query1: 以結關日取已出貨之資料編號
	query1 = f"""
			SELECT ETC, ID FROM EXPORT_SUMMARY
			WHERE ETC >= '{year}0101'
			AND ETC <= '{year}1231'
			AND STATUS = 'SHIPPED'
		"""
	export = pd.read_sql_query(query1, conn)
	export_id = "', '".join(export["ID"].astype(str))
	
	#query2: 透過編號進invoice找淨重
	query2 = f"""
			SELECT N_W, EXPORT_ID FROM INVOICE_SUMMARY
			WHERE EXPORT_ID IN ('{export_id}')
		"""
	invoice = pd.read_sql_query(query2, conn)
	
	conn.close()
	#同一編號會有多個對應項次，將重量加總以利後續合併資料，並將重量(KG)轉換為MT
	invoice = invoice.groupby("EXPORT_ID", as_index=False)["N_W"].sum()
	invoice["ORDER_WEIG_MT"] = invoice["N_W"] / 1000

	#用EXPORT_ID合併query1及query2之資料，以計算本月及上月之接單噸數(保留近兩月避免跨月時前月資料有遺漏)
	#此處資料會比0810Q早更新，本月結果資料重量會較重一些，但為實際資料。上月應相同。
	export["ORDER_WEIG_MT"] = export["ID"].map(invoice.set_index("EXPORT_ID")["ORDER_WEIG_MT"])
	export["ETC"] = pd.to_datetime(export["ETC"], format='%Y%m%d', errors='coerce')
	export["MONTH"] = export["ETC"].dt.strftime('%Y%m')
	have_shipped = export[["MONTH", "ORDER_WEIG_MT"]].groupby("MONTH").sum()
	
	return have_shipped

#將所有結果儲存至檔案
def WRITE_FILE():

	output_path =  fr"Z:\業務部\業務一課\H-訂單\3. 接單統計\{year}年度\螺絲訂單統計-{year}年度.xlsx"
	wb = load_workbook(output_path)

	 
	for month, df in monthly_dfs.items():
		month_name = month.replace("/", "-")
		sheet_name = f"{month_name}月接單明細"

		# If sheet already exists, remove it first to avoid duplication
		if sheet_name in wb.sheetnames:
			del wb[sheet_name]

		# Add new sheet
		ws = wb.create_sheet(title=sheet_name)

		# Write DataFrame to sheet row by row
		for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
			for c_idx, value in enumerate(row, start=1):
				ws.cell(row=r_idx, column=c_idx, value=value)


	ws = wb[f"螺絲接單暨出貨狀況表-{year}"]
	if not unconfirm.empty:
		value = unconfirm['ORDER_WEIG_MT'].sum()
		ws["C61"] = value if pd.notna(value) else 0
	else:
		ws["C61"] = 0

	order_weig_list = month_summerize["ORDER_WEIG_MT"].tolist()
	expect_ship_list = expect_ship["ORDER_WEIG_MT"].tolist()
	have_shipped_list = have_shipped["ORDER_WEIG_MT"].tolist()

	start_row = 6

	for i, weig in enumerate(order_weig_list):
		ws.cell(row=start_row + i, column=3, value=float(weig))

	for i, expect in enumerate(expect_ship_list):
		ws.cell(row=start_row + i, column=8, value=float(expect))

	for i, shipped in enumerate(have_shipped_list):
		ws.cell(row=start_row +1+ i, column=9, value=float(shipped))

	wb.save(output_path)
	
	return output_path

def CAPTURE_RESULT(excel_path, sheet_name, cell_range, image_path):

	excel = win32.Dispatch('Excel.Application')
	excel.Visible = False  # Keep Excel hidden
	
	wb = excel.Workbooks.Open(excel_path, UpdateLinks=False, ReadOnly=True)
	
	ws = wb.Sheets(sheet_name)
	ws.Activate()
	ws.Range(cell_range).Select()

	time.sleep(3)  # Give Excel time to render the range

	ws.Range(cell_range).CopyPicture(Format=-4147)

	# Create a large chart to paste the picture
	left, top, width, height = 50, 50, 1200, 800
	chart = ws.ChartObjects().Add(left, top, width, height)
	chart.Activate()
	chart.Chart.Paste()
	chart.Chart.Export(image_path)

	wb.Close(SaveChanges=False)
	excel.Quit()

	return image_path


#自動發送檔案
def SEND_MAIL(to, cc, subject, today, attachment_path, inline_image_path):
	now = datetime.now()
	outlook = win32.Dispatch('Outlook.Application')
	mail = outlook.CreateItem(0)  # 0: olMailItem

	mail.To = "; ".join(to)
	mail.Subject = subject
	mail.CC = cc

	image_cid = "image"
	html_body = f"""
	<html>
	<body>
		<p>Dear all,</p>
		<p>螺絲接單統計至 {today}，謝謝。</p>
		<p><img src="cid:{image_cid}"></p>
		<p>Best regards</p>
		<p>統計時間:{now}</p>
	</body>
	</html>
	"""

	mail.HTMLBody = html_body
	
	mail.Attachments.Add(Source=attachment_path)
	image_attachment = mail.Attachments.Add(Source=inline_image_path)
	image_attachment.PropertyAccessor.SetProperty(
		"http://schemas.microsoft.com/mapi/proptag/0x3712001F", image_cid
	)

	mail.Send()  # mail.Display() or mail.Send()


if __name__ == "__main__":

	def log(message):
		timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
		log_message = f"【LOG - {timestamp}】{message}"
		print(log_message)
		with open(r"C:\Users\sales6\Desktop\task_log_sumscrew.txt", "a", encoding="utf-8") as f:
			f.write(log_message + "\n")

	log("螺絲接單統計腳本啟動中...")

	year, last_year, today, next_year = GET_DATE()
	unconfirm, monthly_dfs, month_summerize, expect_ship = ERP_CONNECT(year, last_year)
	have_shipped = DB_CONNECT(year, last_year)
	
	output_path = WRITE_FILE()
	sheet_name = f"螺絲接單暨出貨狀況表-{year}"
	image_output = r"C:\Users\sales6\Desktop\output.png"

	image_path = CAPTURE_RESULT(
		excel_path=output_path,
		sheet_name=sheet_name,
		cell_range="A1:I62",
		image_path=image_output
	)
	
	SEND_MAIL(
		to=["amy@soonport.com", "gavin@soonport.com", "raymond@soonport.com", "rita_yang@soonport.com"],
		cc="bu1s@soonport.com",
		subject=f"螺絲接單統計-{today}更新",
		today=today,
		attachment_path=output_path,
		inline_image_path=image_path
	)
	
	log("螺絲接單統計腳本執行完成，郵件已發送。")

	