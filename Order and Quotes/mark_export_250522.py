import pandas as pd
import os
import cx_Oracle
import re
from openpyxl import Workbook
from openpyxl.styles import Font


class MARK_EXPORT():
    def __init__(self):
        self.ALL = None
        self.New_Item = None
        self.Order_Number = None
        self.SC_Number = None
        self.Customer_Code = None
        self.PATH = input("Enter the order path: ")
        self.ORDER_INFO()
        self.FETCH_DATA()

    def ORDER_INFO(self):
        for file in os.listdir(self.PATH):
            if "0210M" in file:
                path = os.path.join(self.PATH, file)
        self.ALL = pd.read_excel(path,dtype={'客戶產品代號(P/N)': str})

        self.New_Item = self.ALL[self.ALL["初次下單"] == "Y"]
        self.New_Item = self.New_Item.reset_index(drop=True)
        

    def FETCH_DATA(self):
        First_Item = self.ALL["客戶產品代號(P/N)"].iloc[0]

        if self.New_Item.empty:
            print("無初次下單 不需麥頭")
            exit()

        # Define connection details
        dsn = cx_Oracle.makedsn(
            host = "192.168.1.242",      # Replace with your host IP or domain
            port = 1526,                # Replace with your port
            service_name = "sperpdb"  # Replace with your service name
        )

        # Establish the connection
        connection = cx_Oracle.connect(
            user = "spselect",         # Replace with your username
            password = "select",     # Replace with your password
            dsn = dsn
        )
        query = f"SELECT SC_NO,CST_REFE_NO,ORD_CST_NO FROM V_SCH0200Q_ORD WHERE CST_PART_NO = '{First_Item}' ORDER BY SC_NO DESC "
        df = pd.read_sql_query(query, connection)
        self.Order_Number = df["CST_REFE_NO"].iloc[0]
        self.SC_Number = df["SC_NO"].iloc[0]
        self.Customer_Code = df["ORD_CST_NO"].iloc[0]
        connection.close()
        
        return self.Order_Number, self.SC_Number, self.Customer_Code

    def FILL_EXCEL_EXPORT(self):
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "工作指示單附件"
        ws["A1"].font = Font(size=20, bold=True, underline="single")
        ws["A2"] = "包裝單與棧板嘜頭英文敘述"
        ws["A2"].font = Font(size=14)
        ws["A3"] = "客戶代號"
        ws["A3"].font = Font(size=14)
        ws["B3"] = self.Customer_Code
        ws["B3"].font = Font(size=14)
        ws["A4"] = "SC"
        ws["A4"].font = Font(size=14)
        ws["B4"] = self.SC_Number
        ws["B4"].font = Font(size=14)
        ws["A5"] = "客戶訂單號碼"
        ws["A5"].font = Font(size=14)
        ws["B5"] = self.Order_Number
        ws["B5"].font = Font(size=14)

        ws["A7"] = "項次"
        ws["B7"] = "客戶產品代號(P/N)"
        ws["C7"] = "產品說明(中)"
        ws["D7"] = "客戶指定產品名稱"
        ws["E7"] = "客戶指定電鍍名稱"
        ws["F7"] = "客戶指定產品名稱(嘜頭)"
        ws["G7"] = "客戶指定電鍍名稱(嘜頭)"

        header_row = 8  # Start writing data from row 8 (just below the headers)

        for i, row in self.New_Item.iterrows():
            ws[f"A{header_row + i}"] = row["項次"]
            ws[f"B{header_row + i}"] = row["客戶產品代號(P/N)"]
            ws[f"C{header_row + i}"] = row["產品說明(中)"]
            ws[f"D{header_row + i}"] = row["客戶指定產品名稱"]
            ws[f"E{header_row + i}"] = row["客戶指定電鍍名稱"]
            ws[f"F{header_row + i}"] = row["客戶指定產品名稱(嘜頭)"]
            ws[f"G{header_row + i}"] = row["客戶指定電鍍名稱(嘜頭)"]

        # Define the path for saving the new Excel file
        Attach = os.path.join(self.PATH, "工作指示單附件-嘜頭.xlsx")

        # Save the workbook to the specified path
        wb.save(Attach)

if __name__ == "__main__":
    bot = MARK_EXPORT()  # Instantiate the class
    bot.FILL_EXCEL_EXPORT()

    print("嘜頭儲存至路徑")
        
        
    

    